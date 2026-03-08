"""Local-file retrieval layer for workpapers.

This acts as the first source adapter for the MVP. Today it reads from a local
folder. Later the same interface can be backed by MCP servers or enterprise
connectors.
"""
from __future__ import annotations

import os
from functools import lru_cache
from pathlib import Path
from typing import Annotated

from langchain_core.documents import Document
from langchain_core.tools import tool
from langchain_community.document_loaders import CSVLoader, PyMuPDFLoader, TextLoader
from langchain_openai import OpenAIEmbeddings
from langchain_qdrant import QdrantVectorStore
from langchain_text_splitters import RecursiveCharacterTextSplitter
from openpyxl import load_workbook


def _safe_read_text(path: Path) -> list[Document]:
    try:
        return TextLoader(str(path), encoding="utf-8").load()
    except UnicodeDecodeError:
        return TextLoader(str(path), encoding="latin-1").load()


def _load_spreadsheet(path: Path) -> list[Document]:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    docs: list[Document] = []

    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            normalized = ["" if cell is None else str(cell) for cell in row]
            if any(cell.strip() for cell in normalized):
                rows.append(" | ".join(normalized))

        if rows:
            docs.append(
                Document(
                    page_content=(
                        f"Workbook: {path.name}\n"
                        f"Sheet: {sheet.title}\n"
                        + "\n".join(rows[:400])
                    ),
                    metadata={
                        "source": str(path),
                        "sheet": sheet.title,
                        "file_type": "spreadsheet",
                    },
                )
            )
    return docs


def load_source_documents(data_dir: str) -> list[Document]:
    root = Path(data_dir)
    if not root.exists():
        return []

    docs: list[Document] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue

        suffix = path.suffix.lower()
        try:
            if suffix == ".pdf":
                docs.extend(PyMuPDFLoader(str(path)).load())
            elif suffix in {".md", ".txt"}:
                docs.extend(_safe_read_text(path))
            elif suffix == ".csv":
                docs.extend(CSVLoader(file_path=str(path), encoding="utf-8").load())
            elif suffix in {".xlsx", ".xlsm"}:
                docs.extend(_load_spreadsheet(path))
        except Exception as exc:
            docs.append(
                Document(
                    page_content=f"Could not parse {path.name}: {exc}",
                    metadata={"source": str(path), "parse_error": True},
                )
            )

    return docs


@lru_cache(maxsize=1)
def _get_vector_store() -> QdrantVectorStore | None:
    data_dir = os.environ.get("RAG_DATA_DIR", "data")
    documents = load_source_documents(data_dir)
    if not documents:
        return None

    splitter = RecursiveCharacterTextSplitter(
        chunk_size=int(os.environ.get("RAG_CHUNK_SIZE", "900")),
        chunk_overlap=int(os.environ.get("RAG_CHUNK_OVERLAP", "150")),
    )
    chunks = splitter.split_documents(documents)
    if not chunks:
        return None

    embeddings = OpenAIEmbeddings(
        model=os.environ.get("OPENAI_EMBED_MODEL", "text-embedding-3-small")
    )
    return QdrantVectorStore.from_documents(
        documents=chunks,
        embedding=embeddings,
        location=":memory:",
        collection_name="audit_workpapers",
    )


def search_workpapers(query: str, *, k: int = 4) -> list[Document]:
    store = _get_vector_store()
    if store is None:
        return []

    retriever = store.as_retriever(search_kwargs={"k": k})
    return retriever.invoke(query)


def collect_workpaper_evidence(queries: list[str], *, k: int = 4) -> list[dict[str, str]]:
    results: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    counter = 1

    for query in queries:
        docs = search_workpapers(query, k=k)
        for doc in docs:
            metadata = doc.metadata or {}
            source = Path(metadata.get("source", "unknown")).name
            location_bits = []
            if metadata.get("page") is not None:
                location_bits.append(f"page {metadata['page']}")
            if metadata.get("sheet"):
                location_bits.append(f"sheet {metadata['sheet']}")
            location = ", ".join(location_bits) if location_bits else "document body"
            excerpt = " ".join(doc.page_content.split())[:900]
            dedupe_key = (source, location, excerpt)
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            results.append(
                {
                    "citation": f"S{counter}",
                    "source": source,
                    "location": location,
                    "query": query,
                    "excerpt": excerpt,
                }
            )
            counter += 1

    return results


def describe_indexed_sources() -> str:
    data_dir = os.environ.get("RAG_DATA_DIR", "data")
    root = Path(data_dir)
    if not root.exists():
        return "No source directory found."

    files = [path.name for path in sorted(root.rglob("*")) if path.is_file()]
    if not files:
        return "No source files found."

    return "Indexed source files:\n- " + "\n- ".join(files)


@tool
def retrieve_workpaper_context(
    query: Annotated[str, "Grounded query over uploaded workpapers"],
    top_k: Annotated[int, "Number of evidence chunks to return"] = 4,
) -> str:
    """Retrieve grounded evidence from the indexed workpaper folder."""
    items = collect_workpaper_evidence([query], k=top_k)
    if not items:
        return "No indexed workpaper context was found for that query."

    blocks = []
    for item in items:
        blocks.append(
            f"[{item['citation']}] Source: {item['source']} ({item['location']})\n"
            f"Query: {item['query']}\n"
            f"Excerpt: {item['excerpt']}"
        )
    return "\n\n".join(blocks)
